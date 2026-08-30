#!/usr/bin/env python3
"""
調教師評価Excelジェネレーター

JRA-VAN SQLite DB (jravan.db) から指定年の調教師別
勝利数・重賞勝利数・ランキング順位を集計してExcelに出力する。

Usage:
    python -m app.tools.generate_trainer_eval \\
        --db shared/jravan.db \\
        --year 2025 \\
        --output trainer_eval_2025.xlsx
"""

import argparse
import sqlite3
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# グレードコード (コード表2003): 重賞と判定するコード
# A=G1, B=G2, C=G3（平地）/ F=J-G1, G=J-G2, H=J-G3（障害）/ D=グレードなし重賞
GRADE_CODES_G123 = ("A", "B", "C")       # 平地G1/G2/G3
GRADE_CODES_ALL  = ("A", "B", "C", "D")  # 平地重賞全て（グレードなし含む）

WINS_QUERY = """
SELECT
    r.trainer_code,
    COALESCE(t.trainer_name, r.trainer_name_short) AS trainer_name,
    r.east_west_code,
    CASE r.east_west_code
        WHEN '1' THEN '美浦'
        WHEN '2' THEN '栗東'
        ELSE 'その他'
    END AS affiliation,
    COUNT(*)                                                       AS total_entries,
    SUM(CASE WHEN r.finish_pos = 1 THEN 1 ELSE 0 END)             AS wins,
    SUM(CASE WHEN r.finish_pos = 1 AND ra.grade_code = 'A'
             THEN 1 ELSE 0 END)                                    AS g1_wins,
    SUM(CASE WHEN r.finish_pos = 1 AND ra.grade_code = 'B'
             THEN 1 ELSE 0 END)                                    AS g2_wins,
    SUM(CASE WHEN r.finish_pos = 1 AND ra.grade_code = 'C'
             THEN 1 ELSE 0 END)                                    AS g3_wins,
    SUM(CASE WHEN r.finish_pos = 1 AND ra.grade_code IN ('A','B','C')
             THEN 1 ELSE 0 END)                                    AS graded_wins,
    SUM(CASE WHEN r.finish_pos = 1 AND ra.grade_code IN ('A','B','C','D')
             THEN 1 ELSE 0 END)                                    AS all_graded_wins
FROM horse_race_results r
LEFT JOIN races ra
    ON  r.kaisai_year = ra.kaisai_year
    AND r.kaisai_date = ra.kaisai_date
    AND r.venue_code  = ra.venue_code
    AND r.kai         = ra.kai
    AND r.nichi       = ra.nichi
    AND r.race_no     = ra.race_no
LEFT JOIN (
    SELECT trainer_code, trainer_name
    FROM trainer_yearly_stats
    WHERE period_type = 'current' AND period_year = ?
    GROUP BY trainer_code
) t ON r.trainer_code = t.trainer_code
WHERE r.kaisai_year = ?
  AND r.east_west_code IN ('1', '2')
  AND (r.abnormal_code IS NULL OR r.abnormal_code IN ('', '0'))
GROUP BY r.trainer_code, r.east_west_code
HAVING wins > 0
ORDER BY wins DESC
"""


def fetch_stats(db_path: str, year: int) -> pd.DataFrame:
    conn = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True)
    try:
        df = pd.read_sql_query(WINS_QUERY, conn, params=(year, str(year)))
    finally:
        conn.close()
    return df


def add_ranking(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["wins_rank"]        = df["wins"].rank(method="min", ascending=False).astype(int)
    df["graded_wins_rank"] = df["graded_wins"].rank(method="min", ascending=False).astype(int)
    return df


# カラム定義: (ヘッダー表示名, DataFrameカラム名 or None=計算列, 列幅)
COLUMNS = [
    ("勝利数順位",       "wins_rank",        8),
    ("調教師コード",     "trainer_code",    10),
    ("調教師名",         "trainer_name",    12),
    ("所属",             "affiliation",      6),
    ("勝利数",           "wins",             7),
    ("重賞勝利(G1-3)",   "graded_wins",     12),
    ("G1",               "g1_wins",          5),
    ("G2",               "g2_wins",          5),
    ("G3",               "g3_wins",          5),
    ("重賞勝利(全)",     "all_graded_wins", 12),
    ("出走数",           "total_entries",    8),
    ("勝率",             None,               7),  # 計算列: wins / total_entries
]

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F497D")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Arial", size=10)
DATA_FONT   = Font(name="Arial", size=10)
ALT_FILL    = PatternFill(fill_type="solid", fgColor="DDEBF7")
G1_FONT     = Font(name="Arial", size=10, bold=True, color="C00000")
CENTER      = Alignment(horizontal="center", vertical="center")
LEFT        = Alignment(horizontal="left",   vertical="center")
RIGHT       = Alignment(horizontal="right",  vertical="center")
THIN_BORDER = Border(
    bottom=Side(border_style="thin", color="BFBFBF"),
)


def write_excel(df: pd.DataFrame, output_path: str, year: int) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}年調教師評価"
    ws.freeze_panes = "A2"

    # ヘッダー行
    for col_idx, (header, _, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 18

    # データ行
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        is_alt = (row_idx % 2 == 0)
        wins_val    = int(row.get("wins", 0) or 0)
        entries_val = int(row.get("total_entries", 0) or 0)

        for col_idx, (_, field, _) in enumerate(COLUMNS, 1):
            if field is None:
                # 勝率: 計算列
                value = f"{wins_val/entries_val:.1%}" if entries_val > 0 else "0.0%"
                align = RIGHT
            elif field in ("trainer_name", "affiliation"):
                value = row.get(field, "")
                align = LEFT
            else:
                value = row.get(field)
                if value is not None:
                    try:
                        value = int(value)
                    except (ValueError, TypeError):
                        pass
                align = CENTER if field in ("trainer_code", "east_west_code") else RIGHT

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = align
            cell.border    = THIN_BORDER

            # G1勝利数を強調
            if field == "g1_wins" and isinstance(value, int) and value > 0:
                cell.font = G1_FONT
            else:
                cell.font = DATA_FONT

            if is_alt:
                cell.fill = ALT_FILL

    # 集計情報をフッターに追記
    footer_row = len(df) + 3
    ws.cell(row=footer_row, column=1, value=f"対象年: {year}年").font = Font(name="Arial", size=9, italic=True)
    ws.cell(row=footer_row, column=3, value=f"調教師数: {len(df)}名").font = Font(name="Arial", size=9, italic=True)
    note = "重賞勝利(G1-3)=平地G1/G2/G3のみ  重賞勝利(全)=グレードなし重賞(D)含む  出典: JRA-VAN"
    ws.cell(row=footer_row + 1, column=1, value=note).font = Font(name="Arial", size=8, italic=True, color="808080")

    wb.save(output_path)
    print(f"Saved: {output_path}  ({len(df)} trainers)")


def main() -> None:
    parser = argparse.ArgumentParser(description="調教師評価Excel生成")
    parser.add_argument("--db",     default="shared/jravan.db", help="SQLiteパス")
    parser.add_argument("--year",   type=int, default=2025,     help="対象年 (default: 2025)")
    parser.add_argument("--output", default=None,               help="出力Excelパス")
    args = parser.parse_args()

    db_path    = Path(args.db)
    output     = args.output or f"trainer_eval_{args.year}.xlsx"

    if not db_path.exists():
        print(f"Error: DB not found: {db_path}")
        raise SystemExit(1)

    print(f"Querying {db_path} for year {args.year}...")
    df = fetch_stats(str(db_path), args.year)

    if df.empty:
        print(f"Warning: No race result data for {args.year}. "
              "Run JVLinkFetcher --setup --race-from-year <YEAR> on Windows first.")
        raise SystemExit(1)

    df = add_ranking(df)
    write_excel(df, output, args.year)


if __name__ == "__main__":
    main()
