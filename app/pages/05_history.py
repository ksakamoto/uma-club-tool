"""過去データ管理ページ：過去年Excel一括インポート + 活躍状況トラッキング。"""
from __future__ import annotations
import streamlit as st
import pandas as pd
import openpyxl

from app.core.userdata_writer import UserDataWriter
from app.core.db_reader import DBReader

APPLY_COL_HINTS      = ["申込", "応募", "申し込み"]
WON_COL_HINTS        = ["当選", "確定", "当確", "結果"]
HORSE_COL_HINTS      = ["馬名", "募集馬名"]
BANGOU_COL_HINTS     = ["番号"]
WON_VALUES           = {"当選", "当確", "○", "〇", "◎", "✓", "✔", "OK", "ok", "Ok"}
EVAL_VIDEO_OPTIONS   = ["", "◎", "◯", "▲", "△", "×"]
EVAL_OVERALL_OPTIONS = ["", "◎", "◯", "×", "★"]
_HISTORY_BASELINE    = "history_eval_baseline"
_EVAL_MAP = {"A": "◎", "B": "◯", "C": "▲", "*": "△"}


def _find_sheet(names: list[str], hints: list[str]) -> str | None:
    for h in hints:
        for n in names:
            if h in n:
                return n
    return None


def _read_sheet(ws) -> pd.DataFrame:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()
    header = [str(c) if c is not None else "" for c in rows[0]]
    data = [list(r) for r in rows[1:]]
    return pd.DataFrame(data, columns=header)


def _find_col(columns, hints: list[str]) -> str | None:
    for h in hints:
        for c in columns:
            if h in str(c):
                return c
    return None


def _truthy(v) -> bool:
    if v is None:
        return False
    return str(v).strip() not in ("", "0", "False", "なし", "×", "nan")


def _has_value(v) -> bool:
    if v is None:
        return False
    return str(v).strip() not in ("", "nan", "None", "#N/A", "#VALUE!", "#REF!")


def _map_eval(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    if not s or s in ("nan", "None"):
        return None
    # 表記揺れを正規化してから標準記号チェック
    _NORM = {"○": "◯", "▴": "▲", "▵": "▲", "✕": "×", "✖": "×"}
    s = _NORM.get(s, s)
    if s in ("◎", "◯", "▲", "△", "×", "★"):
        return s
    # A/B/C/* 記法 → 標準記号。*A など複合は先頭文字で判定
    return _EVAL_MAP.get(s) or _EVAL_MAP.get(s.lstrip("*")) or None

st.title("📚 過去データ管理")

# --- 接続確認 ---
ud: UserDataWriter | None = st.session_state.get("userdata")
db: DBReader | None = st.session_state.get("db")

if not ud or not ud.is_connected():
    st.warning("データ準備ページで接続テストを先に実行してください。")
    st.stop()

tab_import, tab_tracking = st.tabs(["📥 過去データインポート", "📊 活躍状況"])

# ============================================================
# タブ1: 過去データインポート
# ============================================================
with tab_import:
    st.subheader("過去年度 Excelインポート")
    st.markdown("""
過去年の募集馬Excelファイルをインポートします。
- **「一覧」シート** → 基本情報（馬名・父・母・調教師・牧場・性別・価格）
- **「応募」シート**（あれば）→ 申込・当選フラグを自動マッピング
""")

    col_y, col_f = st.columns([1, 3])
    with col_y:
        import_year = st.number_input("対象年度", min_value=2015, max_value=2035, value=2025, step=1)
    with col_f:
        import_file = st.file_uploader("Excelファイルを選択", type=["xlsx"], key="history_import")

    if import_file:
        try:
            wb = openpyxl.load_workbook(import_file, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
        except Exception as e:
            st.error(f"Excelを開けませんでした: {e}")
            st.stop()

        st.caption(f"シート一覧: {sheet_names}")

        # --- 一覧シートを読む ---
        ichiran_name = _find_sheet(sheet_names, ["一覧", "sheet1", "Sheet1"])
        if not ichiran_name:
            st.error("「一覧」シートが見つかりません。シート名を確認してください。")
            st.stop()

        from app.core.csv_parser import COLUMN_ALIASES, _normalize
        df_list = _read_sheet(wb[ichiran_name])
        df_list = df_list.rename(columns=COLUMN_ALIASES)
        df_list.columns = [c.replace("\n", "").strip() if isinstance(c, str) else c for c in df_list.columns]
        for col in list(df_list.columns):
            if str(col).startswith("動画") and "eval_video" not in df_list.columns:
                df_list = df_list.rename(columns={col: "eval_video"})
            elif str(col).startswith("総合評価") and "eval_overall" not in df_list.columns:
                df_list = df_list.rename(columns={col: "eval_overall"})
        if "horse_name" not in df_list.columns:
            st.error("「一覧」シートに馬名列が見つかりません。")
            st.stop()
        df_list = df_list[df_list["horse_name"].notna() & (df_list["horse_name"].astype(str).str.strip() != "")]
        df_list = _normalize(df_list)

        st.success(f"「{ichiran_name}」シートから {len(df_list)} 頭を読み込みました")
        st.dataframe(df_list[["horse_name"] + [c for c in ["sire_name","dam_name","trainer_name","sex","price_total"] if c in df_list.columns]].head(10), use_container_width=True)

        # --- 応募シートを読む ---
        applied_map: dict[str, dict] = {}
        oubo_name = _find_sheet(sheet_names, ["応募", "申込"])
        if oubo_name:
            df_oubo = _read_sheet(wb[oubo_name])
            horse_col  = _find_col(df_oubo.columns, HORSE_COL_HINTS)
            apply_col  = _find_col(df_oubo.columns, APPLY_COL_HINTS)
            won_col    = _find_col(df_oubo.columns, WON_COL_HINTS)
            bangou_col = _find_col(df_oubo.columns, BANGOU_COL_HINTS)
            if horse_col:
                for _, r in df_oubo.iterrows():
                    if bangou_col and not _has_value(r.get(bangou_col)):
                        continue
                    name = str(r[horse_col]).strip()
                    if not _has_value(name):
                        continue
                    applied_map[name] = {
                        "applied": 1,
                        "won": 1 if won_col and str(r.get(won_col, "")).strip() in WON_VALUES else 0,
                    }
                st.info(f"「{oubo_name}」シートから {len(applied_map)} 頭の申込情報を読み込みました")
                applied_count = sum(1 for v in applied_map.values() if v["applied"])
                won_count = sum(1 for v in applied_map.values() if v["won"])
                st.caption(f"申込済み: {applied_count}頭 / 当選: {won_count}頭")

        # --- インポート実行 ---
        if st.button(f"💾 {import_year}年分をインポート", type="primary"):
            rows = []
            for _, r in df_list.iterrows():
                name = str(r.get("horse_name", "")).strip()
                flags = applied_map.get(name, {})
                rows.append({
                    "horse_name":   name,
                    "sire_name":    r.get("sire_name"),
                    "dam_name":     r.get("dam_name"),
                    "trainer_name": r.get("trainer_name"),
                    "farm_name":    r.get("farm_name"),
                    "sex":          r.get("sex"),
                    "price_total":  r.get("price_total"),
                    "recruit_no":   r.get("recruit_no"),
                    "eval_video":   _map_eval(r.get("eval_video")),
                    "eval_overall": _map_eval(r.get("eval_overall")),
                    "applied":      flags.get("applied", 0),
                    "won":          flags.get("won", 0),
                })
            count, skipped = ud.import_from_excel_rows(import_year, rows)
            st.success(f"{count} 頭をインポートしました（{import_year}年）")
            if skipped:
                st.warning(f"スキップ: {skipped}")

# ============================================================
# タブ2: 活躍状況
# ============================================================
with tab_tracking:
    st.subheader("過去募集馬の活躍状況")

    all_years = ud.get_all_years()
    if not all_years:
        st.info("インポートタブから過去データを取り込むとここに表示されます。")
        st.stop()

    col_fy, col_ff = st.columns([1, 2])
    with col_fy:
        sel_year = st.selectbox("年度", all_years)
    with col_ff:
        eval_filter = st.multiselect(
            "総合評価フィルタ", ["◎", "◯", "×", "★"], default=[],
            help="未選択=全て"
        )

    horses_df = ud.get_horses_by_year(sel_year)
    if horses_df.empty:
        st.info(f"{sel_year}年のデータがありません。")
        st.stop()

    if eval_filter:
        horses_df = horses_df[horses_df["eval_overall"].isin(eval_filter)]

    applied_only = st.checkbox("申込した馬のみ", value=False)
    if applied_only:
        horses_df = horses_df[horses_df["applied"] == 1]

    # JRA-VANから成績を取得（募集名→正式馬名を解決してから検索）
    if db and db.is_connected():
        recruitment_names = horses_df["horse_name"].tolist()
        with st.spinner("JRA-VANから成績を照合中..."):
            resolved_df = db.resolve_recruitment_names(recruitment_names)
            official_names = resolved_df["official_name"].tolist()
            perf_df = db.get_performance_summary(official_names)

        # 正式馬名で結合してから recruitment_name に戻す
        if not perf_df.empty:
            perf_df = perf_df.merge(
                resolved_df.rename(columns={"official_name": "horse_name"}),
                on="horse_name", how="left",
            )
            perf_df["official_name"] = perf_df["horse_name"]
            perf_df["horse_name"] = perf_df["recruitment_name"].fillna(perf_df["horse_name"])
            perf_df = perf_df.drop(columns=["recruitment_name", "resolved"], errors="ignore")
            merged = horses_df.merge(perf_df, on="horse_name", how="left")
        else:
            merged = horses_df.copy()
            for col in ["total_runs", "wins", "top3", "total_prize", "graded_wins", "graded_placed"]:
                merged[col] = None

        # 名前解決できなかった馬を表示
        unresolved = resolved_df[~resolved_df["resolved"]]["recruitment_name"].tolist()
        if unresolved:
            with st.expander(f"⚠️ 正式馬名が見つからなかった馬 ({len(unresolved)}頭)"):
                st.caption("「母名の年」形式でなかった、またはまだデビュー前の可能性があります。")
                st.write(unresolved)
    else:
        merged = horses_df.copy()
        st.warning("JRA-VAN DBに接続されていないため成績を取得できません。")

    display_cols = [
        "recruit_no", "horse_name", "official_name", "eval_overall", "eval_video",
        "applied", "won",
        "total_runs", "wins", "total_prize", "graded_wins",
    ]
    display_cols = [c for c in display_cols if c in merged.columns]
    display_df = merged[display_cols].copy()

    _baseline_key = f"{_HISTORY_BASELINE}_{sel_year}"

    def _snap(df: pd.DataFrame) -> dict:
        return {
            row["horse_name"]: {
                "eval_video": row.get("eval_video"),
                "eval_overall": row.get("eval_overall"),
            }
            for _, row in df.iterrows()
        }

    _before = st.session_state.get(_baseline_key)

    col_config = {
        "recruit_no":    st.column_config.NumberColumn("No.", disabled=True, width="small", format="%d"),
        "horse_name":    st.column_config.TextColumn("募集時馬名", disabled=True, width=130),
        "official_name": st.column_config.TextColumn("現在の馬名", disabled=True, width=130),
        "trainer_name":  st.column_config.TextColumn("調教師", disabled=True),
        "eval_video":    st.column_config.SelectboxColumn("動画/現場", options=EVAL_VIDEO_OPTIONS, width="small"),
        "eval_overall":  st.column_config.SelectboxColumn("総合評価", options=EVAL_OVERALL_OPTIONS, width="small"),
        "applied":       st.column_config.CheckboxColumn("申込", disabled=True, width="small"),
        "won":           st.column_config.CheckboxColumn("当選", disabled=True, width="small"),
        "total_runs":    st.column_config.NumberColumn("出走", disabled=True, format="%d"),
        "wins":          st.column_config.NumberColumn("勝利", disabled=True, format="%d"),
        "total_prize":   st.column_config.NumberColumn("賞金(万)", disabled=True, format="%.1f"),
        "graded_wins":   st.column_config.NumberColumn("重賞勝", disabled=True, format="%d"),
    }

    edited = st.data_editor(
        display_df,
        column_config=col_config,
        use_container_width=True,
        hide_index=True,
        height=500,
        key=f"history_editor_{sel_year}",
    )

    if ud:
        _after = _snap(edited)
        if _before is None:
            st.session_state[_baseline_key] = _after
        else:
            changed = {name: vals for name, vals in _after.items() if _before.get(name) != vals}
            if changed:
                _change_df = pd.DataFrame([{"horse_name": k, **v} for k, v in changed.items()])
                ud.update_evals_bulk(sel_year, _change_df)
                st.toast("保存しました ✓")
            new_baseline = {**_before}
            new_baseline.update(_after)
            st.session_state[_baseline_key] = new_baseline

    st.caption(f"表示: {len(merged)} 頭")
